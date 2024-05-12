import argparse
from openpyxl  import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd
from fuzzywuzzy import process,fuzz
from deepdiff import DeepDiff

FUZZY_TOLERANCE = 80

def verify(team, jersey, name, id, target:pd.DataFrame):
    return process.extractBests(name.upper().strip(),target['NAME'],scorer=fuzz.token_sort_ratio,score_cutoff=FUZZY_TOLERANCE-0.01)

def load_source(source:str):
    df =  pd.read_csv(source, index_col =0)
    df['NAME'] = df['FIRST_NAME'].str.upper() + ' ' + df['LAST_NAME'].str.upper()
    return df.copy()

def get_roster(worksheet:Worksheet):
    a,b = zip(*[
        ((jersey.value,a_name.value,a_id.value),(jersey.value,b_name.value,b_id.value)) 
        for jersey,a_name,a_id,b_name,b_id in worksheet['A2:E101']
        ]
    )
    a_players = [('A',jersey,name.strip().upper(),id)for jersey,name,id in a if name and name.strip()]
    b_players = [('B',jersey,name.strip().upper(),id) for jersey,name,id in b if name and name.strip()]
    return pd.DataFrame(a_players+b_players,columns = ['Team','Jersey', 'Name','ID']).copy()


def main(location:str , source:str, strict_match:bool = False, overwrite:bool = False):
    wb = load_workbook(location,read_only= not overwrite)
    rosters = wb['ROSTERS']
    extr = get_roster(rosters)
    source = load_source(source)
    print(source)
    found = True
    ambig = False
    not_found = []
    ambig_list = []
    new_df = []
    old = [tuple(a) for a in extr.itertuples()]
    for player in old:
        index,team, jersey, name, id = player
        # print(f'{team}-{jersey}-{name}-{id}')
        matches = verify(team,jersey,name,id,source)
        if len(matches) == 1: 
            verified_name,match_pct,verified_id = matches[0]
            if match_pct == 100: 
                # Perfect Match
                pass
            else: 
                print(f'Matching {team}{jersey}:{name} -> {verified_name}: Confidence {match_pct}') 
            new_df.append((index,team,jersey,verified_name,verified_id))
        elif matches:
            verified_name,match_pct,verified_id = matches[0]
            
            if match_pct != 100:
                ambig = True
                print('Ambiguous match (>1 high probability match) found: ')
                print(f'Matching {team}{jersey}:{name} -> :') 
                ambig_list.append(name)
                print(matches)
            else:
                # print('Selecting the 100% Match: ')
                print(f'Matching {team}{jersey}:{name} -> {verified_name}: Confidence {match_pct}') 
            new_df.append((index,team,jersey,verified_name,verified_id))
        else:
            found = False
            print(f'Who the hell is {name}')
            not_found.append(name)
            
    nl = "\n"
    # print(DeepDiff(old,new_df))
    if not found: 
        print(f'ERROR: Players {nl.join(not_found)} not found')
        return
    elif ambig and strict_match:
        print(f'ERROR: Ambiguous matches needed for {nl.join(ambig_list)}')
        return
    else:
        print(f'Validated roster document @ {location}')


    if overwrite:
        print(f'Overwriting rosters with validated names and IDS @ {location}')
        for _,team, jersey, name, id in new_df:
            delt = 2 if team == 'A' else 4
            rosters.cell(int(jersey)+2,delt,name)
            rosters.cell(int(jersey)+2,delt+1,id)
        # wb['ROSTERS'] = rosters
        wb.save(location)
        

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('statsheet_location')
    parser.add_argument('--source', default='upstream')
    parser.add_argument('--strict-match', action = 'store_true')
    parser.add_argument('--overwrite', action = 'store_true')
    args = parser.parse_args()
    main(args.statsheet_location, args.source, args.strict_match, args.overwrite)