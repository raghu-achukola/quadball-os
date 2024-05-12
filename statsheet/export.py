from openpyxl import load_workbook
from statsheet.gen_possessions import * 
from statsheet.statsheet import * 


class StatsheetHandler:
    def __init__(self, file_location:str, *, verify_upstream = False, read_only = True):
        self.wb = load_workbook(file_location, read_only=read_only)
        self._validate_wb_structure()
        self._read_sheets()

        self.metadata = self._read_metadata()
        self.possessions = self._read_possessions()
        self.rosters = self._read_rosters()
        print(self.occurrences)

    def _validate_wb_structure(self): 
        important_sheets = {'POSSESSIONS','METADATA','ROSTERS'}
        assert set(self.wb.sheetnames).intersection(important_sheets) == important_sheets, f"""
            Workbook's sheets {self.wb.sheetnames} do not include one or more of necessary sheets:
            {important_sheets} 
        """
    
    def _read_sheets(self):
        self.possession_sheet = self.wb['POSSESSIONS']
        self.metadata_sheet = self.wb['METADATA']
        self.rosters_sheet = self.wb['ROSTERS']
    
    def _read_metadata(self) -> StatSheetMetadata:
        metadata_params = {}
        for row in self.metadata_sheet.iter_rows():
            if row[0].value:
                metadata_params[row[0].value.lower().replace(' ','_')] =  row[1].value
            else:
                break
        return StatSheetMetadata(**metadata_params)

    def _read_possessions(self, starter_cell = (7,1)) -> list[StatSheetPossession]: 
        possession_list = []
        self.occurrences = {}
        possessions = self.possession_sheet
        for i,poss in enumerate(gen_statsheet_values(possessions, starter_cell= possessions.cell(*starter_cell))):
            poss = StatSheetPossession.from_values(*poss)
            possession_list.append(poss)
            extras = poss.extras if poss.extras else []
            primary  = poss.primary if poss.primary else []
            secondary = poss.secondary if poss.secondary else []
            for e,extra in enumerate(extras):
                if extra.player:
                    self.occurrences[extra.player]  = self.occurrences.get(extra.player,{})
                    self.occurrences[extra.player]['extras'] = self.occurrences[extra.player].get('extras',[])+[(i+1,e+1)]
            for _,player in enumerate(primary):
                if player:
                    self.occurrences[player] = self.occurrences.get(player,{})
                    self.occurrences[player]['primary'] = self.occurrences[player].get('primary',[]) + [(i+1,0)]
            for _,player in enumerate(secondary):
                if player:
                    self.occurrences[player] = self.occurrences.get(player,{})
                    self.occurrences[player]['secondary'] = self.occurrences[player].get('primary',[]) + [(i+1,0)]                    


        return possession_list
    
    def _read_rosters(self) -> StatSheetRoster:
        a,b = zip(*[
            ((jersey.value,a_name.value,a_id.value),(jersey.value,b_name.value,b_id.value)) 
            for jersey,a_name,a_id,b_name,b_id in self.rosters_sheet['A2:E101']
            ]
        )
        a_players = {
            ('A',jersey):{
                'name':name.strip().upper(),
                'id':id
            }
            for jersey,name,id in a if name and name.strip()
        }
        b_players = {
            ('B',jersey):{
                'name':name.strip().upper(),
                'id':id
            } 
            for jersey,name,id in b if name and name.strip()
        }
        return StatSheetRoster.from_dict( a_players | b_players )

    def validate_coverage(self) -> list:
        for player, occurences in self.occurrences:
            if player not in self.rosters.player_dict:
                print(f'PLAYER NOT FOUND IN ROSTER, {player}')

    def get_hydrated(self) -> list: 
        team_lookup = {'A':self.metadata.team_a_name,'B':self.metadata.team_b_name}
        hydrated = [possession.copy() for possession in self.possessions]
        for poss in hydrated:
            
            poss.offense = team_lookup[poss.offense]
            if poss.primary:
                for i,pl in enumerate(poss.primary):
                    team, player = pl
                    poss.primary[i] = f"{team_lookup[team]}-{self.rosters[pl]['name']}"
            if poss.secondary:
                for i,pl in enumerate(poss.secondary):
                    team, player = pl
                    poss.secondary[i] = f"{team_lookup[team]}-{self.rosters[pl]['name']}"
            if poss.extras:
                for i,ex in enumerate(poss.extras):
                    if ex.player:
                        team,number = ex.player
                        poss.extras[i].player = f"{team_lookup[team]}-{self.rosters[ex.player]['name']}"
                    if ex.team:
                        poss.extras[i].team = f"{team_lookup[team]}"
                    
        return hydrated
    