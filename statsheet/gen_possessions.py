import argparse
from openpyxl  import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
import pandas as pd
from typing import Generator
from enum import Enum
import json

from statsheet.statsheet import *



def get_cell_group_values(ws:Worksheet ,starter_row,starter_col) -> list:
    """
        Given a worksheet and a starting row and column. this function 
        will output the values left -> right then up -> down of the 3x3 rectangle
        starting at the starter cell. This if placed correctly represents a 
        'block' in the statsheet representing a possesssion. 
        The default on the statsheet should be A6 aka (6,1)
    """
    return [
        ws.cell(starter_row + r, starter_col +c).value
        for r in range(3)
        for c in range(3)
    ]

def gen_statsheet_values(ws:Worksheet, starter_cell:Cell) -> Generator[list,None,None]:
    """
        Generator method
        Bounces from possession block to possession block yielding the 
        statsheet values in the 3x3 possession block each time
    """ 
    first = True
    data_found = False
    while True: 
        print('HO')
        starter_cell = starter_cell.offset(0,0 if first else 3)
        first = False
        # NOTE: If we ever replace the A or B, this will BREAK
        if starter_cell.offset(1,1).value not in ('A','B'): 
            print('it not in A!')
            # We ran out of the columns in this row
            # Start over in the next row 
            # NOTE: we are assuming the cell starts over at 1 (no padding)
            starter_cell = ws.cell(starter_cell.row +3, 1)
            if starter_cell.offset(1,1).value not in ('A','B'): 
                # We ran out of rows
                break
        data =  get_cell_group_values( ws, starter_cell.row, starter_cell.column)
        # print(data)
        if not any(data[5:]):
            if data_found:
                # we ran out of rows
                break
            else:
                continue
        else:
            data_found = True
        d2 =  [None if cell is None else str(cell) for cell in data]
        # print(d2)
        yield d2


def get_roster(roster_sheet:Worksheet):
    a,b = zip(*[
        ((jersey.value,a_name.value,a_id.value),(jersey.value,b_name.value,b_id.value)) 
        for jersey,a_name,a_id,b_name,b_id in roster_sheet['A2:E101']
        ]
    )
    a_players = [('A',jersey,name.strip().upper(),id)for jersey,name,id in a if name and name.strip()]
    b_players = [('B',jersey,name.strip().upper(),id) for jersey,name,id in b if name and name.strip()]
    return pd.DataFrame(a_players+b_players,columns = ['Team','Jersey', 'Name','ID']).copy()





def main(location:str, out:str, translate_roster:str = None):
    wb = load_workbook(location)
    possessions = wb['POSSESSIONS']
    rosters = wb['ROSTERS']
    metadata = wb['METADATA']
    translator = {}
    if translate_roster:
        translator = {
            (v['Team'],str(v['Jersey'])):(v['Name'] if translate_roster == 'name' else v['ID'])
            for _, v in get_roster(rosters).T.to_dict().items()
        }
        team_a,team_b = (4,6) if translate_roster == 'name' else (5,7)
        translator|={'A': metadata.cell(team_a,2).value,'B':metadata.cell(team_b,2).value}
    print(translator)
    sv = gen_statsheet_values(possessions, starter_cell= possessions.cell(6,1))

    poss_list = []
    for poss in sv:
        print(f'POSS: {poss}')
        possession = StatSheetPossession.from_values(*poss)
        if translator:
            poss_list.append(possession.translate(translator))
        else:
            poss_list.append(possession.to_dict())
    if out:
        with open(out,'w') as f:
            # print(possessions.to_json())
            f.write(json.dumps(poss_list))
    else:
        print(poss_list)




    




if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('statsheet_location')
    parser.add_argument('--out')
    parser.add_argument('--translate-roster')
    args = parser.parse_args()
    tr = args.translate_roster
    out = args.out
    if tr:
        if tr.lower().strip() not in ('name', 'id'):
            exit()
        else:
            tr = tr.lower().strip()
    main(args.statsheet_location, out, tr)